#!/usr/bin/env python3
"""
Evidence & Analysis Workbook — 마크다운 11개 파일 → 제출물 02 xlsx 변환기.

이 폴더의 *.md가 정본이고 xlsx는 파생물이다. md를 고친 뒤 이 스크립트를 다시 돌리면
02_Evidence_Analysis_Workbook.xlsx가 갱신된다.

사용법:  python3 build_workbook_xlsx.py
"""
import re
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).parent
OUT = HERE / "02_Evidence_Analysis_Workbook.xlsx"

# 파일 → 시트명 (엑셀 시트명 31자 제한)
SHEETS = [
    ("01_Five_Forces.md",                 "01_Five_Forces"),
    ("02_Value_Chain.md",                 "02_Value_Chain"),
    ("03_KSF.md",                         "03_KSF"),
    ("04_TAM_SAM_SOM_Segment_Map.md",     "04_TAM_SAM_SOM"),
    ("05_Persona_Spectrum_Journey.md",    "05_Persona_Journey"),
    ("06_Opportunity_Score.md",           "06_Opportunity_Score"),
    ("07_JTBD.md",                        "07_JTBD"),
    ("08_Value_Proposition.md",           "08_Value_Proposition"),
    ("09_PRD_Requirement_Trace.md",       "09_PRD_Trace"),
    ("10_Evidence_Inference_Assumption.md", "10_Fact_Inference_Assum"),
    ("11_Benchmark_Transfer.md",          "11_Benchmark_Transfer"),
]

# ── 색 ──────────────────────────────────────────────────────────────────
NAVY   = "1F3864"   # 문서 제목
BLUE   = "BBDEFB"   # ## 절
LBLUE  = "E3F2FD"   # ### 소절
GREY   = "D9D9D9"   # 표 헤더
YELLOW = "FFF9C4"   # 인용(주의·판단)
MONO   = "F2F2F2"   # 산식 블록

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean(text: str) -> str:
    """인라인 마크다운을 셀 텍스트로 편다."""
    t = text.strip()
    t = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", t)   # 링크 → 라벨만
    t = t.replace("**", "").replace("~~", "")
    t = re.sub(r"`([^`]*)`", r"\1", t)
    t = t.replace("<br/>", " / ").replace("<br>", " / ")
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def split_row(line: str):
    """표 행을 셀로 나눈다. 백틱(`) 안의 | 는 구분자가 아니라 내용이다."""
    raw = line.strip().strip("|")
    cells, buf, in_code = [], [], False
    for ch in raw:
        if ch == "`":
            in_code = not in_code
            buf.append(ch)
        elif ch == "|" and not in_code:
            cells.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    cells.append("".join(buf))
    return [clean(c) for c in cells]


def is_sep(line: str) -> bool:
    return bool(re.match(r"^\|[\s:|-]+\|?\s*$", line.strip()))


def parse(md: str):
    """(kind, payload) 블록 목록으로 자른다."""
    blocks, lines, i = [], md.split("\n"), 0
    while i < len(lines):
        ln = lines[i]
        s = ln.strip()

        if s.startswith("```"):                      # 코드/다이어그램
            lang = s[3:].strip().lower()
            body, i = [], i + 1
            while i < len(lines) and not lines[i].strip().startswith("```"):
                body.append(lines[i])
                i += 1
            i += 1
            if lang == "mermaid":
                blocks.append(("mermaid", None))
            else:
                blocks.append(("code", body))
            continue

        if s.startswith("|"):                        # 표
            rows, i = [], i
            while i < len(lines) and lines[i].strip().startswith("|"):
                if not is_sep(lines[i]):
                    rows.append(split_row(lines[i]))
                i += 1
            if rows:
                blocks.append(("table", rows))
            continue

        if s.startswith("#"):
            lvl = len(s) - len(s.lstrip("#"))
            blocks.append((f"h{min(lvl,3)}", clean(s.lstrip('#'))))
        elif s.startswith(">"):
            blocks.append(("quote", clean(s.lstrip('>'))))
        elif s.startswith("---"):
            blocks.append(("rule", None))
        elif s:
            blocks.append(("para", clean(s)))
        i += 1
    return blocks


def cjk_len(text: str) -> int:
    """한글·한자·전각은 2폭, 그 외 1폭으로 센다."""
    n = 0
    for ch in text:
        n += 2 if ord(ch) > 0x2E80 else 1
    return n


def write_sheet(ws, blocks, title):
    r = 1
    # 열 너비는 '표 셀'만 근거로 잡는다 — 산문 길이에 끌려가면 표가 과하게 넓어진다.
    samples = {}

    def note(col, text):
        samples.setdefault(col, []).append(cjk_len(text))

    # 문서 제목
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 26
    r += 2

    for kind, payload in blocks:
        if kind == "h1":
            continue                                  # 제목은 위에서 이미 씀
        if kind == "rule":
            continue

        if kind in ("h2", "h3"):
            c = ws.cell(row=r, column=1, value=payload)
            c.font = Font(bold=True, size=12 if kind == "h2" else 11)
            c.fill = PatternFill("solid", fgColor=BLUE if kind == "h2" else LBLUE)
            c.alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = 22 if kind == "h2" else 18
            r += 2 if kind == "h2" else 1
            continue

        if kind == "table":
            head, *body = payload
            for j, v in enumerate(head, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(bold=True, size=10)
                c.fill = PatternFill("solid", fgColor=GREY)
                c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
                c.border = BORDER
                note(j, v)
            r += 1
            for row in body:
                for j, v in enumerate(row, start=1):
                    c = ws.cell(row=r, column=j, value=v)
                    c.font = Font(size=10)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.border = BORDER
                    note(j, v)
                r += 1
            r += 1
            continue

        if kind == "code":
            for ln in payload:
                c = ws.cell(row=r, column=1, value=ln)
                c.font = Font(name="Courier New", size=10)
                c.fill = PatternFill("solid", fgColor=MONO)
                c.alignment = Alignment(vertical="top")
                r += 1
            r += 1
            continue

        if kind == "mermaid":
            c = ws.cell(row=r, column=1, value="[다이어그램 — 원문 .md 참조]")
            c.font = Font(size=9, italic=True, color="808080")
            r += 2
            continue

        if kind == "quote":
            c = ws.cell(row=r, column=1, value="※ " + payload)
            c.font = Font(size=10, italic=True)
            c.fill = PatternFill("solid", fgColor=YELLOW)
            c.alignment = Alignment(vertical="top")
            r += 2
            continue

        # para
        c = ws.cell(row=r, column=1, value=payload)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top")
        r += 1

    # 열 너비 — 표 셀 길이의 80백분위를 쓰고 상·하한으로 조인다
    for j, lens in samples.items():
        lens = sorted(lens)
        pct = lens[min(int(len(lens) * 0.8), len(lens) - 1)]
        est = max(pct, min(lens[-1], pct + 8))          # 최장값이 튀면 일부만 반영
        est = min(max(int(est * 0.62) + 4, 11), 46)     # 폭 단위 환산 후 11~46자로 제한
        if j == 1:
            est = max(est, 20)                           # 첫 열은 항목명이 들어와 좁으면 읽기 어렵다
        ws.column_dimensions[get_column_letter(j)].width = est
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def write_index(ws, summary_rows):
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 16

    c = ws.cell(row=1, column=1, value="CardFit — Evidence & Analysis Workbook (제출물 02)")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 30

    meta = [
        "Master Deck(제출물 01)이 결론을, 이 Workbook이 계산과 출처를 담는다.",
        "정본은 저장소의 'Evidence & Analysis Workbook/*.md' 11개 파일이며 이 xlsx는 파생물이다 "
        "(build_workbook_xlsx.py로 재생성).",
        "표기 — 🔵 Fact(1차 출처 확인) · ⚪ Inference(추론) · 🟡 Assumption(팀 가정) · 🟠 미확인 / "
        "충족도 ✔ 충족 · ◐ 부분 · ✗ 미충족 · ？ 확인 불가",
        "확인일은 별도 표기가 없으면 2026-08-20. 이 Workbook 작성일은 2026-08-21.",
        "8필드(방법론.md 8장) — 분석 대상 / 구분 / 관찰·주장 / 출처·확인일 / 근거·추론 과정 / "
        "신뢰도 / 영향 / 검증 계획. 각 시트 4절이 이 형식을 따른다.",
        "시트 공통 구조 — 0 최종 판단 · 1 상세표 · 2 산식·계산 · 3 출처·확인일 · "
        "4 Fact/Inference/Assumption · 5 사례 비교 · 6 전이 분석 · 7 추가 검증과제",
    ]
    r = 3
    for m in meta:
        c = ws.cell(row=r, column=1, value=m)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    head = ["탭", "방법론", "정본(근거 원문)", "최종 판단", "상태"]
    for j, v in enumerate(head, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=GREY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    r += 1
    for row in summary_rows:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[r].height = 42
        r += 1

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


SUMMARY = [
    ["01", "Porter's Five Forces", "윤정/확정본/01",
     "산업 매력도 중간. 구매자 힘·대체재 위협이 높고 공급자 의존도도 높지만, 3조건 동시 충족 경쟁자 0곳이라는 공백이 진입을 정당화한다 🔵", "✅"],
    ["02", "가치사슬", "윤정/확정본/02",
     "차별화 구간은 체인의 중간 두 단계(제약 최적화·결제수단 배분). 경쟁자 체인에 아예 없다 🔵. 스코프 경계(해지 실행 지원 없음)의 정의처", "✅"],
    ["03", "Top 5 KSF", "윤정/확정본/03",
     "1·2는 Table Stakes(기존 사업자 보유), 3(손실 가시화+근거 투명성)이 유일하게 이길 수 있는 지점 ⚪", "✅"],
    ["04", "TAM-SAM-SOM & Segment Map", "윤정/확정본/04",
     "TAM은 누적 가입 1억 7,734만 건으로만 표기(사람 수 환산 안 함) → SAM 12개월 내 소비변화자 → SOM 혼인 Beachhead", "🔶 SAM·SOM 실수치 🟠"],
    ["05", "페르소나 스펙트럼·여정지도", "가람/확정본/05",
     "12명 중 4명이 7단계를 완주하지 못하고 ③온보딩·⑥실행이 가장 취약. ③은 기능(F-11)으로, ⑥은 측정(F-13)으로만 대응", "✅ (12명 🟡 가상)"],
    ["06", "기회점수 AOS·DOS", "가람/확정본/06",
     "A와 C가 사실상 공동 1위(SOM 3.04 vs 3.00 / SAM 2.40 vs 2.32). F는 AOS 동급이나 DOS 니치", "🔶 점수 전량 🟡"],
    ["07", "JTBD", "가람/확정본/07",
     "Job 단위 선언 6개 + 경계 3개. 가장 큰 가정이 '입력할까'에서 '혜택 보상이 입력 노동을 정당화하나'로 좁혀졌다", "🔶 모의 인터뷰"],
    ["08", "Value Proposition Sheet", "효진/확정본/08",
     "Fit ✅3·🔶3. 공동 1위 기회 C에 기능이 없고, 기능이 아니라 측정(F-13)으로 대응. North Star 조합안 선택률 ≥40% 🟡", "✅"],
    ["09", "PRD & Requirement Trace", "효진/확정본/09",
     "핵심 기능은 F-04(Net Benefit 게이팅) 하나. Guardrail 5개 중 하나라도 넘으면 중단. E2에서 20% 미달이면 피벗", "✅ / D1~D5 🔴"],
    ["10", "Fact/Inference/Assumption 집계", "정본 없음 — 이 Workbook이 정본",
     "🔵13 · ⚪6 · 🟡8 · 🟠22. E2 Concierge Test 하나가 최대 리스크 4건을 동시에 해소한다", "✅ 신규 작성"],
    ["11", "벤치마크·메커니즘 전이", "윤정/확정본/11 + 윤정/0820",
     "뱅크샐러드(진단)·핀트(게이팅)·토스(전달) 채택, 2사를 서로 다른 이유로 제외. 핀트의 '자동 실행'을 뺀 것이 핵심 변형", "✅"],
]


def main():
    wb = Workbook()
    write_index(wb.active, SUMMARY)
    wb.active.title = "00_Index"

    for fname, sheetname in SHEETS:
        path = HERE / fname
        md = path.read_text(encoding="utf-8")
        blocks = parse(md)
        title = next((p for k, p in blocks if k == "h1"), sheetname)
        ws = wb.create_sheet(title=sheetname[:31])
        write_sheet(ws, blocks, title)
        print(f"  {sheetname:26s} ← {fname}  ({ws.max_row}행 × {ws.max_column}열)")

    wb.save(OUT)
    print(f"\n저장: {OUT.name}  (시트 {len(wb.sheetnames)}개)")


if __name__ == "__main__":
    main()
