#!/usr/bin/env python3
"""
제출물 03 — Decision & AI Usage Log xlsx 변환기.

원본(정본):
  ../decision-log/decision-log.md                  전 결정 이력 (날짜별 표)
  ../ai-usage-log/ai-usage-log.md                  3인 AI 활용 로그
  ../ai-usage-log/methodology.md                   3원칙·컬럼 가이드
  ../효진/0820/final/03_Decision_AI_Usage_Log.md    결론을 바꾼 것만 골라낸 큐레이션본

xlsx는 파생물이다. 원본 md를 고친 뒤 이 스크립트를 다시 돌려 갱신한다.
표 파싱은 build_workbook_xlsx.py의 함수를 그대로 쓴다.

사용법:  python3 build_decision_log_xlsx.py
"""
import re
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from build_workbook_xlsx import (
    clean, split_row, is_sep, cjk_len,
    NAVY, BLUE, GREY, YELLOW,
)

HERE = pathlib.Path(__file__).parent
ROOT = HERE.parent
OUT = HERE / "03_Decision_AI_Usage_Log.xlsx"

SRC_DECISION = ROOT / "decision-log" / "decision-log.md"
SRC_AILOG = ROOT / "ai-usage-log" / "ai-usage-log.md"
SRC_CURATED = ROOT / "효진" / "0820" / "final" / "03_Decision_AI_Usage_Log.md"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ANOMALIES = []      # 원본 데이터 이상을 숨기지 않고 00_Index에 적는다


# ── 원본 파싱 ────────────────────────────────────────────────────────────
def tables_by_heading(md: str, level: str = "## "):
    """{소제목: [표행, ...]} — 같은 소제목 아래 여러 표는 이어 붙인다."""
    out, cur = {}, None
    lines = md.split("\n")
    i = 0
    while i < len(lines):
        ln = lines[i]
        if ln.startswith(level) and not ln.startswith(level + "#"):
            cur = clean(ln[len(level):])
            out.setdefault(cur, [])
            i += 1
            continue
        if ln.strip().startswith("|"):
            rows = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                if not is_sep(lines[i]):
                    rows.append(split_row(lines[i]))
                i += 1
            if cur is not None:
                out[cur].append(rows)
            continue
        i += 1
    return out


def flatten(md: str, key_name: str, expect: int):
    """날짜/담당자별 표를 한 장으로 편다. 첫 열에 그 구분값을 넣는다."""
    grouped = tables_by_heading(md)
    header, body = None, []
    for group, tables in grouped.items():
        if group in ("작성 가이드",):
            continue
        for rows in tables:
            if not rows:
                continue
            # 원본에 빈 줄이 끼면 한 표가 여러 조각으로 잘린다. 첫 행이 이미 확정된
            # 헤더와 같을 때만 헤더로 보고, 다르면 전부 데이터로 취급한다 —
            # 그러지 않으면 잘린 조각의 첫 행(실제 결정 1건)이 헤더로 먹혀 사라진다.
            if header is None:
                head, *data = rows
                header = [key_name] + head
            elif rows[0] == header[1:]:
                data = rows[1:]
            else:
                data = rows
                ANOMALIES.append(
                    f"{key_name} '{group}' — 원본 표가 빈 줄로 잘려 조각 {len(rows)}행이 "
                    f"머리글 없이 이어졌다(원본 md에서도 표가 끊겨 보인다). 데이터로 살렸다: "
                    f"{(rows[0][1] if len(rows[0]) > 1 else '')}/{(rows[0][2] if len(rows[0]) > 2 else '')[:28]}…"
                )
            for r in data:
                if len(r) != expect:
                    ANOMALIES.append(
                        f"{key_name} '{group}' — 열 {len(r)}개(표준 {expect}개): "
                        f"{(r[0] if r else '')[:24]}… → 부족분은 빈 칸으로 채웠다"
                    )
                r = (r + [""] * expect)[:expect]
                body.append([group] + r)
    return header, body


def curated_blocks(md: str):
    """### D-NN 아래의 '| 항목 | 내용 |' 2열 표를 한 행으로 전치한다."""
    out = []
    chunks = re.split(r"\n### ", md)
    for ch in chunks[1:]:
        title_line, *rest = ch.split("\n")
        title = clean(title_line)
        if not re.match(r"^D-\d+", title):
            continue
        fields, notes = {}, []
        for ln in rest:
            s = ln.strip()
            if s.startswith("###") or s.startswith("## ") or s.startswith("# "):
                break
            if s.startswith("|") and not is_sep(s):
                cells = split_row(s)
                if len(cells) >= 2 and cells[0] not in ("항목",):
                    fields[cells[0]] = cells[1]
            elif s.startswith(">"):
                notes.append(clean(s.lstrip(">")))
        did, _, name = title.partition(".")
        out.append((did.strip(), name.strip(), fields, " / ".join(notes)))
    return out


def find_table(md: str, heading_prefix: str):
    """'## 제목' 또는 '### 제목' 바로 아래 첫 표를 돌려준다."""
    lines = md.split("\n")
    for i, ln in enumerate(lines):
        if ln.lstrip("#").strip().startswith(heading_prefix) and ln.startswith("#"):
            rows, j = [], i + 1
            while j < len(lines) and not lines[j].strip().startswith("|"):
                if lines[j].startswith("#"):
                    break
                j += 1
            while j < len(lines) and lines[j].strip().startswith("|"):
                if not is_sep(lines[j]):
                    rows.append(split_row(lines[j]))
                j += 1
            if rows:
                return rows
    return []


# ── 쓰기 ────────────────────────────────────────────────────────────────
def write_table(ws, header, body, title, subtitle="", autofilter=False):
    r = 1
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[r].height = 26
    r += 1
    if subtitle:
        c = ws.cell(row=r, column=1, value=subtitle)
        c.font = Font(size=10, italic=True)
        r += 1
    r += 1

    head_row = r
    for j, v in enumerate(header, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=GREY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    r += 1

    samples = {}
    for row in body:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            samples.setdefault(j, []).append(cjk_len(str(v)))
        r += 1

    for j, lens in samples.items():
        lens = sorted(lens)
        pct = lens[min(int(len(lens) * 0.8), len(lens) - 1)]
        est = max(pct, min(lens[-1], pct + 8))
        est = min(max(int(est * 0.62) + 4, 11), 46)
        ws.column_dimensions[get_column_letter(j)].width = est
    for j in range(1, len(header) + 1):
        if j not in samples:
            ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = ws.cell(row=head_row + 1, column=1).coordinate
    if autofilter and body:
        ws.auto_filter.ref = (
            f"A{head_row}:{get_column_letter(len(header))}{head_row + len(body)}"
        )
    ws.sheet_view.showGridLines = False
    return r


def write_index(ws, sheet_rows, counts):
    for col, w in zip("ABCD", (26, 46, 60, 14)):
        ws.column_dimensions[col].width = w

    c = ws.cell(row=1, column=1, value="CardFit — Decision & AI Usage Log (제출물 03)")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 30

    meta = [
        "Part A는 프로젝트 방향을 실제로 바꾼 결정, Part B는 AI 활용과 그 검증 기록이다.",
        f"기간 2026.08.14 ~ 08.21 · 팀 김윤정(기준서비스·시스템설계) · 김가람(문제·메커니즘 리서치) · 이효진(팀장·PRD·검증)",
        f"수록 건수 — 주요 결정 {counts['major']}건 / 전체 결정 로그 {counts['all']}건 / "
        f"판정 철회 {counts['retract']}건 / 미결 {counts['open']}건 / AI 활용 로그 {counts['ai']}건",
        "정본은 원본 md 4개이고 이 xlsx는 파생물이다 — build_decision_log_xlsx.py로 재생성한다.",
        "기존 행은 수정하지 않는다. 내용이 바뀌면 새 행으로 남겨 이력을 보존한다(decision-log·ai-usage-log 공통 원칙).",
    ]
    r = 3
    for m in meta:
        c = ws.cell(row=r, column=1, value=m)
        c.font = Font(size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    def block(title, header, rows):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor=BLUE)
        ws.row_dimensions[r].height = 22
        r += 1
        for j, v in enumerate(header, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=GREY)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
        r += 1
        for row in rows:
            for j, v in enumerate(row, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(size=10)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = BORDER
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    block("시트 구성", ["시트", "내용", "원본", "행"], sheet_rows)

    block("원본 문서", ["경로", "내용", "", ""], [
        ["decision-log/decision-log.md", "전 결정 이력 — 날짜·시간·담당·후보·근거·결정·변경이유·재검토조건", "", ""],
        ["ai-usage-log/ai-usage-log.md", "3인 AI 활용 로그 전체", "", ""],
        ["ai-usage-log/methodology.md", "3원칙·컬럼 가이드·운영 규칙", "", ""],
        ["효진/0820/final/03_Decision_AI_Usage_Log.md", "결론을 바꾼 것만 골라낸 큐레이션본 (Part A·B의 출처)", "", ""],
    ])

    notes = [[a, "", "", ""] for a in ANOMALIES] + [[
        "A3 미결 O-02(GR4 금지어 사전)는 큐레이션본에 '🟠 미작성'으로 남아 있으나 "
        "2026-08-21에 작성 완료됐다(효진/0820/final/06_금지어_사전.md). "
        "로그 원본은 고치지 않고 이 주석으로 남긴다 — 잔여 과제는 '문구 외부화'다.", "", "", ""]]
    block("원본 데이터 주의", ["내용", "", "", ""], notes)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def main():
    dec_md = SRC_DECISION.read_text(encoding="utf-8")
    ai_md = SRC_AILOG.read_text(encoding="utf-8")
    cur_md = SRC_CURATED.read_text(encoding="utf-8")

    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "00_Index"
    sheet_rows = []

    def add(name, header, body, title, subtitle="", autofilter=False, src=""):
        ws = wb.create_sheet(title=name[:31])
        write_table(ws, header, body, title, subtitle, autofilter)
        sheet_rows.append([name, title.split("—")[-1].strip() or title, src, str(len(body))])
        print(f"  {name:24s} {len(body):3d}행 x {len(header)}열")
        return ws

    # ── Part A ──────────────────────────────────────────────────────────
    blocks = curated_blocks(cur_md)
    # 9건 중 1~2건에만 나타나는 필드(이전 결정·문제·변경 이유·파생 결정·파급)는
    # 열로 따로 두면 대부분 빈 칸이 된다. 라벨을 붙여 한 열로 합친다.
    MAIN = ["날짜·담당", "검토한 후보", "결정", "근거", "재검토 조건"]
    # MAIN에 없는 필드는 카드마다 이름이 달라(상황·딜레마·해소·성과·정본 선정 근거 등)
    # 열로 고정할 수 없다. 문서에 나온 순서대로 라벨을 붙여 한 열에 모아 흘려보낸다 —
    # 고정 목록으로 두면 목록에 없는 필드가 조용히 사라진다.
    header_a1 = ["ID", "제목", "날짜·담당", "검토한 후보", "결정", "근거",
                 "배경·변경·파급", "재검토 조건", "메모"]
    body_a1 = []
    for did, name, f, note in blocks:
        merged = " / ".join(f"[{k}] {v}" for k, v in f.items()
                            if k not in MAIN and v)
        body_a1.append([
            did, name, f.get("날짜·담당", ""), f.get("검토한 후보", ""),
            f.get("결정", ""), f.get("근거", ""), merged,
            f.get("재검토 조건", ""), note,
        ])
    extra = sorted({k for _, _, f, _ in blocks for k in f} - set(MAIN))
    print(f"     · 배경·변경·파급 열에 합친 가변 필드 {len(extra)}종: {', '.join(extra)}")
    add("A1_주요결정", header_a1, body_a1,
        "Part A-1 · 방향을 바꾼 주요 결정",
        "큐레이션본 A-1의 D-NN 카드를 한 행으로 전치했다. 전수 이력은 A4 시트에 있다.",
        src="효진/0820/final/03")

    rows = find_table(cur_md, "A-2.")
    add("A2_판정철회", rows[0], rows[1:],
        "Part A-2 · 검증으로 뒤집은 판정",
        "AI 산출물을 그대로 쓰지 않았다는 증적. R-01은 자체 판정을 스스로 철회한 기록이다.",
        src="효진/0820/final/03")

    rows = find_table(cur_md, "A-3.")
    add("A3_미결사항", rows[0], rows[1:],
        "Part A-3 · 남아 있는 미결 사항",
        "🔴 착수 전 선결 · 🟠 확인 필요 · 🟡 실험 대기 · ⬜ 미정. O-02는 00_Index 주의 참조.",
        src="효진/0820/final/03")

    header, body = flatten(dec_md, "날짜", expect=7)
    add("A4_전체결정로그", header, body,
        "Part A-4 · 전체 결정 로그 (전수)",
        "decision-log.md의 날짜별 표를 한 장으로 폈다. 첫 행에서 필터·정렬이 된다.",
        autofilter=True, src="decision-log/decision-log.md")

    # ── Part B ──────────────────────────────────────────────────────────
    rows = find_table(cur_md, "B-1.")
    add("B1_활용3원칙", rows[0], rows[1:],
        "Part B-1 · AI 활용 3원칙",
        '로그의 "잘되지 않은 점" 칸이 비어 있으면 검증을 안 했다는 신호로 보고 반드시 채우기로 했다.',
        src="ai-usage-log/methodology.md")

    body_b2, header_b2 = [], None
    for who in ("김윤정", "김가람", "이효진"):
        rows = find_table(cur_md, who)
        if not rows:
            continue
        if header_b2 is None:
            header_b2 = ["담당자"] + rows[0]
        for r_ in rows[1:]:
            body_b2.append([who] + (r_ + [""] * len(rows[0]))[:len(rows[0])])
    add("B2_담당별요약", header_b2, body_b2,
        "Part B-2 · 담당별 AI 활용 요약",
        "3인의 표를 담당자 열을 붙여 한 장으로 합쳤다.",
        autofilter=True, src="효진/0820/final/03")

    rows = find_table(cur_md, "B-3.")
    add("B3_AI오류_유형", rows[0], rows[1:],
        "Part B-3 · AI가 실제로 틀린 것 — 유형별",
        "틀린 지점이 거의 전부 여러 출처의 시점·범위·정의를 교차할 때였다. "
        "그래서 검증 규칙이 '출처·확인일·정의 범위를 문서에 강제로 남기는 것'으로 수렴했다.",
        src="효진/0820/final/03")

    rows = find_table(cur_md, "B-4.")
    add("B4_사람만한판단", rows[0], rows[1:],
        "Part B-4 · 사람만 한 판단",
        "AI가 할 수 없었던 이유를 함께 남긴다.",
        src="효진/0820/final/03")

    header, body = flatten(ai_md, "담당자", expect=7)
    add("B5_AI로그_전체", header, body,
        "Part B-5 · AI 활용 로그 (3인 전수)",
        "ai-usage-log.md의 담당자별 표를 한 장으로 폈다. 첫 행에서 필터·정렬이 된다.",
        autofilter=True, src="ai-usage-log/ai-usage-log.md")

    counts = {
        "major": len(body_a1),
        "all": len(wb["A4_전체결정로그"]["A"]) and sum(
            1 for _ in wb["A4_전체결정로그"].iter_rows(min_row=5) if _[0].value),
        "retract": len(find_table(cur_md, "A-2.")) - 1,
        "open": len(find_table(cur_md, "A-3.")) - 1,
        "ai": len(body),
    }
    write_index(index_ws, sheet_rows, counts)

    wb.save(OUT)
    print(f"\n저장: {OUT.name}  (시트 {len(wb.sheetnames)}개)")
    if ANOMALIES:
        print("원본 데이터 이상 — 00_Index에 기록:")
        for a in ANOMALIES:
            print("  -", a)


if __name__ == "__main__":
    main()
